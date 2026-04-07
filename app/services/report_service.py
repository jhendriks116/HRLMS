import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from datetime import date

from fastapi.responses import FileResponse

import app.database as db
from app.services.employee_service import EmployeeService

REPORT_DIR = "app/reports"
os.makedirs(REPORT_DIR, exist_ok=True)

# -----------------------------
# Convert ISO string → date
# -----------------------------
def parse_date(value):
    if isinstance(value, str):
        try:
            return date.fromisoformat(value)
        except:
            return value
    return value


def safe_get(data: dict, key: str):
    return data.get(key, "data unavailable") or "data unavailable"


# =====================================================
# FETCH DATA + FIX DOCUMENT DETECTION + DATE CONVERSION
# =====================================================
def fetch_report_data():
    employees = db.read_all_records("employees")
    leave_requests = db.read_all_records("leave_requests")
    leave_balances = db.read_all_records("leave_balances")

    # -----------------------------
    # Map employee leave balances
    # -----------------------------
    balance_map = {}
    for lb in leave_balances:
        emp_id = lb["employee_id"]

        if emp_id not in balance_map:
            balance_map[emp_id] = {"vacation": None, "sick": None}

        if lb["leave_type"] == "vacation":
            balance_map[emp_id]["vacation"] = lb["remaining_days"]
        elif lb["leave_type"] == "sick":
            balance_map[emp_id]["sick"] = lb["remaining_days"]

    # -----------------------------
    # Enhance employee records
    # -----------------------------
    enhanced_emps = []
    for e in employees:
        e["date_hired"] = parse_date(e.get("date_hired"))

        enhanced_emps.append({
            **e,
            "service_years": EmployeeService.calc_years_of_service(e["date_hired"]),
            "vacation_balance": balance_map.get(e["id"], {}).get("vacation", "data unavailable"),
            "sick_balance": balance_map.get(e["id"], {}).get("sick", "data unavailable"),
        })

    # -----------------------------
    # Enhance leave request records
    # -----------------------------
    enhanced_reqs = []
    for lr in leave_requests:
        lr["start_date"] = parse_date(lr.get("start_date"))
        lr["end_date"] = parse_date(lr.get("end_date"))

        # FIX: detect sick document uploaded properly
        doc_uploaded = False
        doc_link = None

        doc_id = lr.get("sick_document_id")
        if doc_id:
            sick_doc = db.find_one("sick_documents", id=doc_id)
            if sick_doc:
                doc_uploaded = True
                doc_link = sick_doc.get("file_path")  # store the file path

        enhanced_reqs.append({
            **lr,
            "has_document": doc_uploaded,
            "document_link": doc_link
        })

    return enhanced_emps, enhanced_reqs


# ======================================
# EXCEL FORMATTING
# ======================================
def autofit_columns(file_path: str):
    wb = openpyxl.load_workbook(file_path)
    for sheet in wb.worksheets:
        for col_cells in sheet.columns:
            max_length = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            sheet.column_dimensions[col_letter].width = max_length + 2
    wb.save(file_path)


def apply_conditional_formatting(file_path: str):
    wb = openpyxl.load_workbook(file_path)

    for sheet in wb.worksheets:
        header = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2):
            # Vacation warning
            try:
                idx = header.index("Remaining Vacation")
                cell = row[idx]
                if isinstance(cell.value, int) and cell.value <= 3:
                    cell.fill = PatternFill(start_color="FFC7CE", fill_type="solid")
            except:
                pass

            # Sick warning
            try:
                idx = header.index("Remaining Sick")
                cell = row[idx]
                if isinstance(cell.value, int) and cell.value <= 2:
                    cell.fill = PatternFill(start_color="FFEB9C", fill_type="solid")
            except:
                pass

            # Pending highlight
            try:
                idx = header.index("Leave Status")
                cell = row[idx]
                if cell.value == "pending":
                    cell.fill = PatternFill(start_color="FFFF00", fill_type="solid")
            except:
                pass

    wb.save(file_path)


# =====================================================
# EXCEL REPORT GENERATOR
# =====================================================
def generate_excel_report(employees: list, leave_records: list):

    emp_map = {e["id"]: e for e in employees}
    rows = []

    for rec in leave_records:
        emp = emp_map.get(rec.get("employee_id"))
        if not emp:
            continue

        rows.append({

            "Employee ID": safe_get(emp, "id"),
            "Employee Name": safe_get(emp, "name"),
            "Email": safe_get(emp, "email"),
            "Employment Type": safe_get(emp, "type"),

            "Employee Start Date": safe_get(emp, "date_hired"),
            "Length of Service (Years)": safe_get(emp, "service_years"),

            "Leave Type": safe_get(rec, "leave_type"),
            "Leave Status": safe_get(rec, "status"),
            "Start Date": safe_get(rec, "start_date"),
            "End Date": safe_get(rec, "end_date"),

            # FIXED
            "Supporting Document Uploaded": "YES" if rec.get("has_document") else "NO",
            "Document Link": rec.get("document_link") or "None",

            "Remaining Vacation": safe_get(emp, "vacation_balance"),
            "Remaining Sick": safe_get(emp, "sick_balance"),
        })

    if not rows:
        raise ValueError("No leave records found.")

    df = pd.DataFrame(rows)

    # Split by employment type
    fortnightly = df[df["Employment Type"] == "fortnightly"]
    monthly = df[df["Employment Type"] == "monthly"]

    work_path = os.path.join(REPORT_DIR, "combined_leave_report.xlsx")

    with pd.ExcelWriter(work_path, engine="openpyxl") as writer:
        if not fortnightly.empty:
            fortnightly.to_excel(writer, sheet_name="Fortnightly", index=False)
        if not monthly.empty:
            monthly.to_excel(writer, sheet_name="Monthly", index=False)

    autofit_columns(work_path)
    apply_conditional_formatting(work_path)

    return work_path
